#!/usr/bin/env python
"""Prepare and build Mango main-image Excel batches.

This script intentionally does not call an image-generation API. It prepares
reference images and builds the final workbook after generated images named
<local_id>.png have been placed in the generated image directory.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import sys
import urllib.request
from dataclasses import asdict, dataclass
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image


REQUIRED_HEADERS = ("本地ID", "产品标题", "产品图片1")
HEADER_ROW = 2
FINAL_IMAGE_SIZE = (800, 800)
DEFAULT_OUTPUT_ROOT = Path(r"D:\MangoMainImageBatches")
LEGACY_BATCH_SUFFIX = "_main_image_batch"
LATEST_POINTER_SUFFIX = "_latest_batch.json"
FINAL_STATUSES = {"verified"}
OPEN_STATUSES = {"pending", "failed"}
CLAIMED_STATUS = "claimed"
TEMPLATE_STATUS = "template_generated"
ALL_STATUSES = ("pending", CLAIMED_STATUS, "verified", "failed", TEMPLATE_STATUS)


@dataclass
class Item:
    row_number: int
    local_id: str
    title: str
    source_image_url: str
    original_image_path: str
    generated_image_path: str


def clean_cell(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def safe_id(value: str) -> str:
    value = clean_cell(value)
    return re.sub(r"[^0-9A-Za-z_-]", "_", value)


def normalize_run_mode(run_mode: str | None, limit: int | None) -> str:
    if run_mode:
        value = run_mode.lower().strip()
        if value not in ("sample", "full"):
            raise ValueError("run_mode must be 'sample' or 'full'")
        return value
    return "full" if limit is None else "sample"


def requested_limit_value(limit: int | None) -> int:
    return 0 if limit is None else limit


def default_batch_id(input_path: Path) -> str:
    return f"{input_path.stem}{LEGACY_BATCH_SUFFIX}"


def create_batch_id(input_path: Path) -> str:
    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S_%f")
    return f"{input_path.stem}_{timestamp}{LEGACY_BATCH_SUFFIX}"


def latest_pointer_path(input_path: Path, output_root: Path | None = None) -> Path:
    root = output_root if output_root is not None else DEFAULT_OUTPUT_ROOT
    return root / f"{input_path.stem}{LATEST_POINTER_SUFFIX}"


def read_latest_batch_id(input_path: Path, output_root: Path | None = None) -> str | None:
    pointer_path = latest_pointer_path(input_path, output_root)
    pointer = read_json(pointer_path, {})
    if isinstance(pointer, dict) and isinstance(pointer.get("batch_id"), str):
        return pointer["batch_id"]
    return None


def write_latest_pointer(input_path: Path, output_root: Path | None, batch_id: str, paths: dict[str, Path]) -> None:
    pointer_path = latest_pointer_path(input_path, output_root)
    pointer_path.parent.mkdir(parents=True, exist_ok=True)
    write_json(
        pointer_path,
        {
            "input_xlsx": str(input_path.resolve()),
            "batch_id": batch_id,
            "batch_dir": str(paths["base_dir"].resolve()),
            "status_ledger": str(paths["status_path"].resolve()),
            "manifest_path": str(paths["manifest_path"].resolve()),
            "updated_at": utc_now(),
        },
    )


def input_fingerprint(input_path: Path) -> dict[str, object]:
    stat = input_path.stat()
    digest = hashlib.sha256()
    with input_path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return {
        "path": str(input_path.resolve()),
        "size": stat.st_size,
        "mtime_ns": stat.st_mtime_ns,
        "sha256": digest.hexdigest(),
    }


def workbook_paths(
    input_path: Path,
    output_path: Path | None = None,
    output_root: Path | None = None,
    batch_id: str | None = None,
) -> dict[str, Path]:
    root = output_root if output_root is not None else DEFAULT_OUTPUT_ROOT
    resolved_batch_id = batch_id or read_latest_batch_id(input_path, output_root) or default_batch_id(input_path)
    base_dir = root / resolved_batch_id
    originals_dir = base_dir / "original_images"
    generated_dir = base_dir / "generated_images"
    generated_fast_dir = base_dir / "generated_fast"
    staging_dir = base_dir / "staging"
    review_queue_dir = base_dir / "review_queue"
    manifest_path = base_dir / "manifest.json"
    status_path = base_dir / "status_ledger.json"
    log_path = base_dir / "process_log.json"
    if output_path is None:
        output_path = base_dir / f"{input_path.stem}_main_image_output.xlsx"
    return {
        "base_dir": base_dir,
        "batch_id": Path(resolved_batch_id).name,
        "originals_dir": originals_dir,
        "generated_dir": generated_dir,
        "generated_fast_dir": generated_fast_dir,
        "staging_dir": staging_dir,
        "review_queue_dir": review_queue_dir,
        "manifest_path": manifest_path,
        "status_path": status_path,
        "log_path": log_path,
        "output_xlsx": output_path,
    }


def find_headers(ws) -> dict[str, int]:
    headers: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        value = clean_cell(ws.cell(row=HEADER_ROW, column=col).value)
        if value in REQUIRED_HEADERS:
            headers[value] = col
    missing = [name for name in REQUIRED_HEADERS if name not in headers]
    if missing:
        raise ValueError(f"Missing required headers on row {HEADER_ROW}: {', '.join(missing)}")
    return headers


def iter_items(input_path: Path, paths: dict[str, Path]) -> tuple[list[Item], list[dict]]:
    wb = load_workbook(input_path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    headers = find_headers(ws)
    items: list[Item] = []
    skips: list[dict] = []
    seen_ids: set[str] = set()

    for row in range(HEADER_ROW + 1, ws.max_row + 1):
        local_id = clean_cell(ws.cell(row=row, column=headers["本地ID"]).value)
        title = clean_cell(ws.cell(row=row, column=headers["产品标题"]).value)
        image_url = clean_cell(ws.cell(row=row, column=headers["产品图片1"]).value)

        if not (local_id or title or image_url):
            continue
        if not local_id or not title or not image_url:
            skips.append({"row_number": row, "reason": "missing_required_field", "local_id": local_id})
            continue
        normalized_id = safe_id(local_id)
        if not normalized_id:
            skips.append({"row_number": row, "reason": "invalid_local_id", "local_id": local_id})
            continue
        if normalized_id in seen_ids:
            skips.append({"row_number": row, "reason": "duplicate_local_id", "local_id": local_id})
            continue

        seen_ids.add(normalized_id)
        original_path = paths["originals_dir"] / f"{normalized_id}.jpg"
        generated_path = paths["generated_dir"] / f"{normalized_id}.png"
        items.append(
            Item(
                row_number=row,
                local_id=normalized_id,
                title=title,
                source_image_url=image_url,
                original_image_path=str(original_path.resolve()),
                generated_image_path=str(generated_path.resolve()),
            )
        )
    wb.close()
    return items, skips


def select_items(items: list[Item], limit: int | None) -> list[Item]:
    if limit is None:
        return items
    return items[:limit]


def download_image(url: str, out_path: Path) -> tuple[bool, str | None]:
    if out_path.exists() and out_path.stat().st_size > 0:
        return True, None
    request = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    try:
        with urllib.request.urlopen(request, timeout=30) as response:
            data = response.read()
        if not data:
            return False, "empty_response"
        out_path.write_bytes(data)
        return True, None
    except Exception as exc:  # noqa: BLE001 - log exact failure for batch review.
        return False, str(exc)


def write_json(path: Path, payload: object) -> None:
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def utc_now() -> str:
    return datetime.now(timezone.utc).isoformat()


def read_json(path: Path, default: object) -> object:
    if not path.exists():
        return default
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return default


def summarize_ledger(ledger: dict) -> dict[str, int]:
    counts: dict[str, int] = {}
    records = ledger.get("records", {}) if isinstance(ledger, dict) else {}
    for record in records.values():
        status = record.get("status", "pending")
        counts[status] = counts.get(status, 0) + 1
    counts["total"] = len(records)
    counts["unfinished"] = sum(
        count for status, count in counts.items() if status not in FINAL_STATUSES | {"total", "unfinished"}
    )
    return counts


def sync_status_ledger(
    paths: dict[str, Path],
    items: list[Item],
    skips: list[dict],
    *,
    input_path: Path,
    run_mode: str,
    requested_limit: int,
    valid_row_count: int,
    fingerprint: dict[str, object],
) -> dict:
    existing = read_json(paths["status_path"], {})
    previous_records = {}
    if isinstance(existing, dict) and isinstance(existing.get("records"), dict):
        previous_records = existing["records"]

    records = {}
    for item in items:
        previous = previous_records.get(item.local_id, {})
        current_status = previous.get("status", "pending") if isinstance(previous, dict) else "pending"
        if current_status not in ALL_STATUSES:
            current_status = "pending"
        generated_path = Path(item.generated_image_path)
        if current_status == "verified" and not generated_path.exists():
            current_status = "pending"
        if current_status != CLAIMED_STATUS:
            owner = None
            claimed_at = None
            lease_expires_at = None
        else:
            owner = previous.get("owner") if isinstance(previous, dict) else None
            claimed_at = previous.get("claimed_at") if isinstance(previous, dict) else None
            lease_expires_at = previous.get("lease_expires_at") if isinstance(previous, dict) else None

        records[item.local_id] = {
            "row_number": item.row_number,
            "local_id": item.local_id,
            "title": item.title,
            "source_image_url": item.source_image_url,
            "original_image_path": item.original_image_path,
            "generated_image_path": item.generated_image_path,
            "generated_fast_image_path": str((paths["generated_fast_dir"] / f"{item.local_id}.png").resolve()),
            "status": current_status,
            "error": previous.get("error") if isinstance(previous, dict) else None,
            "owner": owner,
            "claimed_at": claimed_at,
            "lease_expires_at": lease_expires_at,
            "updated_at": (previous.get("updated_at") if isinstance(previous, dict) else None) or utc_now(),
        }

    ledger = {
        "batch_id": paths["batch_id"],
        "batch_dir": str(paths["base_dir"].resolve()),
        "input_xlsx": str(input_path.resolve()),
        "input_fingerprint": fingerprint,
        "run_mode": run_mode,
        "requested_limit": requested_limit,
        "valid_row_count": valid_row_count,
        "updated_at": utc_now(),
        "records": records,
        "skips": skips,
    }
    write_json(paths["status_path"], ledger)
    return ledger


def prepare(
    input_path: Path,
    limit: int | None,
    output_path: Path | None,
    output_root: Path | None,
    run_mode: str | None = None,
    batch_id: str | None = None,
    new_task: bool = False,
) -> dict:
    normalized_run_mode = normalize_run_mode(run_mode, limit)
    requested_limit = requested_limit_value(limit)
    resolved_batch_id = batch_id or (create_batch_id(input_path) if new_task else None)
    paths = workbook_paths(input_path, output_path, output_root, resolved_batch_id)
    for key in ("base_dir", "originals_dir", "generated_dir", "generated_fast_dir", "staging_dir", "review_queue_dir"):
        paths[key].mkdir(parents=True, exist_ok=True)

    all_items, skips = iter_items(input_path, paths)
    valid_row_count = len(all_items)
    items = select_items(all_items, limit)
    downloads = []
    for item in items:
        ok, error = download_image(item.source_image_url, Path(item.original_image_path))
        downloads.append({"local_id": item.local_id, "ok": ok, "error": error})
    fingerprint = input_fingerprint(input_path)
    ledger = sync_status_ledger(
        paths,
        items,
        skips,
        input_path=input_path,
        run_mode=normalized_run_mode,
        requested_limit=requested_limit,
        valid_row_count=valid_row_count,
        fingerprint=fingerprint,
    )

    payload = {
        "input_xlsx": str(input_path.resolve()),
        "batch_id": paths["batch_id"],
        "batch_dir": str(paths["base_dir"].resolve()),
        "limit": limit,
        "run_mode": normalized_run_mode,
        "requested_limit": requested_limit,
        "valid_row_count": valid_row_count,
        "input_fingerprint": fingerprint,
        "items": [asdict(item) for item in items],
        "skips": skips,
        "downloads": downloads,
        "originals_dir": str(paths["originals_dir"].resolve()),
        "generated_dir": str(paths["generated_dir"].resolve()),
        "generated_fast_dir": str(paths["generated_fast_dir"].resolve()),
        "staging_dir": str(paths["staging_dir"].resolve()),
        "review_queue_dir": str(paths["review_queue_dir"].resolve()),
        "output_xlsx": str(paths["output_xlsx"].resolve()),
        "status_ledger": str(paths["status_path"].resolve()),
        "status_counts": summarize_ledger(ledger),
    }
    write_json(paths["manifest_path"], payload)
    write_json(paths["log_path"], {"skips": skips, "downloads": downloads})
    write_latest_pointer(input_path, output_root, paths["batch_id"], paths)
    payload["manifest_path"] = str(paths["manifest_path"].resolve())
    payload["log_path"] = str(paths["log_path"].resolve())
    return payload


def set_widths(ws, widths: Iterable[int]) -> None:
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def normalize_generated_image(path: Path) -> tuple[bool, str | None]:
    if not path.exists():
        return False, "missing"
    try:
        image = Image.open(path).convert("RGB")
        if image.size != FINAL_IMAGE_SIZE:
            image = image.resize(FINAL_IMAGE_SIZE, Image.Resampling.LANCZOS)
        image.save(path, "PNG", optimize=True)
        return True, None
    except Exception as exc:  # noqa: BLE001 - log exact failure for batch review.
        return False, str(exc)


def validate_final_image(path: Path) -> tuple[bool, str | None]:
    if not path.exists():
        return False, "missing"
    try:
        with Image.open(path) as image:
            if image.size != FINAL_IMAGE_SIZE:
                return False, f"wrong_size:{image.size[0]}x{image.size[1]}"
            image.verify()
        return True, None
    except Exception as exc:  # noqa: BLE001 - log exact failure for batch review.
        return False, str(exc)


def ledger_needs_refresh(ledger: dict, input_path: Path, limit: int | None, run_mode: str) -> bool:
    if not isinstance(ledger, dict) or "records" not in ledger:
        return True
    if ledger.get("run_mode") != run_mode:
        return True
    if ledger.get("requested_limit") != requested_limit_value(limit):
        return True
    fingerprint = ledger.get("input_fingerprint", {})
    if not isinstance(fingerprint, dict):
        return True
    current = input_fingerprint(input_path)
    return fingerprint.get("sha256") != current["sha256"]


def load_or_prepare(
    input_path: Path,
    limit: int | None,
    output_path: Path | None,
    output_root: Path | None,
    run_mode: str | None = None,
    batch_id: str | None = None,
) -> tuple[dict, dict[str, Path]]:
    normalized_run_mode = normalize_run_mode(run_mode, limit)
    paths = workbook_paths(input_path, output_path, output_root, batch_id)
    if not paths["status_path"].exists() or not paths["manifest_path"].exists():
        prepare(input_path, limit, output_path, output_root, normalized_run_mode, paths["batch_id"])
    ledger = read_json(paths["status_path"], {})
    if ledger_needs_refresh(ledger, input_path, limit, normalized_run_mode):
        prepare(input_path, limit, output_path, output_root, normalized_run_mode, paths["batch_id"])
        ledger = read_json(paths["status_path"], {})
    return ledger if isinstance(ledger, dict) else {}, paths


def parse_timestamp(value: str | None) -> datetime | None:
    if not value:
        return None
    try:
        return datetime.fromisoformat(value)
    except ValueError:
        return None


def is_claim_expired(record: dict, now: datetime | None = None) -> bool:
    if record.get("status") != CLAIMED_STATUS:
        return False
    expires_at = parse_timestamp(record.get("lease_expires_at"))
    if expires_at is None:
        return True
    return expires_at <= (now or datetime.now(timezone.utc))


def reset_claim(record: dict) -> None:
    record["status"] = "pending"
    record["owner"] = None
    record["claimed_at"] = None
    record["lease_expires_at"] = None
    record["updated_at"] = utc_now()


def recycle_expired_claims(ledger: dict) -> int:
    records = ledger.get("records", {})
    recycled = 0
    for record in records.values():
        if is_claim_expired(record):
            reset_claim(record)
            recycled += 1
    if recycled:
        ledger["updated_at"] = utc_now()
    return recycled


def unfinished_records(records: dict) -> list[dict]:
    return [record for record in records.values() if record.get("status") not in FINAL_STATUSES]


def status(
    input_path: Path,
    limit: int | None,
    output_path: Path | None,
    output_root: Path | None,
    run_mode: str | None = None,
    batch_id: str | None = None,
) -> dict:
    ledger, paths = load_or_prepare(input_path, limit, output_path, output_root, run_mode, batch_id)
    records = ledger.get("records", {})
    unfinished = unfinished_records(records)
    return {
        "status_ledger": str(paths["status_path"].resolve()),
        "manifest_path": str(paths["manifest_path"].resolve()),
        "output_xlsx": str(paths["output_xlsx"].resolve()),
        "batch_id": paths["batch_id"],
        "batch_dir": str(paths["base_dir"].resolve()),
        "run_mode": ledger.get("run_mode"),
        "requested_limit": ledger.get("requested_limit"),
        "valid_row_count": ledger.get("valid_row_count"),
        "counts": summarize_ledger(ledger),
        "next_unfinished": unfinished[:10],
    }


def next_items(
    input_path: Path,
    limit: int | None,
    output_path: Path | None,
    output_root: Path | None,
    count: int,
    run_mode: str | None = None,
    batch_id: str | None = None,
) -> dict:
    ledger, paths = load_or_prepare(input_path, limit, output_path, output_root, run_mode, batch_id)
    records = ledger.get("records", {})
    unfinished = unfinished_records(records)
    return {
        "status_ledger": str(paths["status_path"].resolve()),
        "batch_id": paths["batch_id"],
        "count": min(count, len(unfinished)),
        "items": unfinished[:count],
    }


def mark_item(
    input_path: Path,
    limit: int | None,
    output_path: Path | None,
    output_root: Path | None,
    local_id: str,
    item_status: str,
    error: str | None,
    owner: str | None = None,
    run_mode: str | None = None,
    batch_id: str | None = None,
) -> dict:
    ledger, paths = load_or_prepare(input_path, limit, output_path, output_root, run_mode, batch_id)
    records = ledger.get("records", {})
    if local_id not in records:
        raise SystemExit(f"Unknown local_id in status ledger: {local_id}")
    record = records[local_id]
    if owner and record.get("owner") not in (owner, None):
        raise SystemExit(f"Cannot mark row owned by {record.get('owner')}: {local_id}")
    if item_status == "verified":
        if owner and record.get("status") != CLAIMED_STATUS:
            raise SystemExit(f"Cannot mark verified; row is not claimed by {owner}: {local_id}")
        ok, normalize_error = normalize_generated_image(Path(record["generated_image_path"]))
        if not ok:
            raise SystemExit(f"Cannot mark verified; generated image is not valid: {normalize_error}")
    record["status"] = item_status
    record["error"] = error
    if item_status != CLAIMED_STATUS:
        record["owner"] = None
        record["claimed_at"] = None
        record["lease_expires_at"] = None
    record["updated_at"] = utc_now()
    ledger["updated_at"] = utc_now()
    write_json(paths["status_path"], ledger)
    return {
        "status_ledger": str(paths["status_path"].resolve()),
        "batch_id": paths["batch_id"],
        "local_id": local_id,
        "status": item_status,
        "counts": summarize_ledger(ledger),
    }


def claim_items(
    input_path: Path,
    limit: int | None,
    output_path: Path | None,
    output_root: Path | None,
    count: int,
    owner: str,
    lease_minutes: int,
    run_mode: str | None = None,
    batch_id: str | None = None,
) -> dict:
    if not owner:
        raise SystemExit("--owner is required for claim")
    ledger, paths = load_or_prepare(input_path, limit, output_path, output_root, run_mode, batch_id)
    recycled = recycle_expired_claims(ledger)
    records = ledger.get("records", {})
    now = datetime.now(timezone.utc)
    expires_at = now + timedelta(minutes=lease_minutes)
    claimable = [
        record for record in sorted(records.values(), key=lambda item: item.get("row_number", 0))
        if record.get("status", "pending") in OPEN_STATUSES
    ]
    claimed = []
    for record in claimable[:count]:
        record["status"] = CLAIMED_STATUS
        record["owner"] = owner
        record["claimed_at"] = now.isoformat()
        record["lease_expires_at"] = expires_at.isoformat()
        record["error"] = None
        record["updated_at"] = utc_now()
        claimed.append(record)
    if claimed or recycled:
        ledger["updated_at"] = utc_now()
        write_json(paths["status_path"], ledger)
    return {
        "status_ledger": str(paths["status_path"].resolve()),
        "batch_id": paths["batch_id"],
        "owner": owner,
        "lease_minutes": lease_minutes,
        "recycled_expired_claims": recycled,
        "count": len(claimed),
        "items": claimed,
        "counts": summarize_ledger(ledger),
    }


def release_items(
    input_path: Path,
    limit: int | None,
    output_path: Path | None,
    output_root: Path | None,
    owner: str,
    local_ids: list[str] | None = None,
    run_mode: str | None = None,
    batch_id: str | None = None,
) -> dict:
    if not owner:
        raise SystemExit("--owner is required for release")
    ledger, paths = load_or_prepare(input_path, limit, output_path, output_root, run_mode, batch_id)
    records = ledger.get("records", {})
    selected = set(local_ids or [])
    released = 0
    for record in records.values():
        if record.get("status") != CLAIMED_STATUS or record.get("owner") != owner:
            continue
        if selected and record.get("local_id") not in selected:
            continue
        reset_claim(record)
        released += 1
    if released:
        ledger["updated_at"] = utc_now()
        write_json(paths["status_path"], ledger)
    return {
        "status_ledger": str(paths["status_path"].resolve()),
        "batch_id": paths["batch_id"],
        "owner": owner,
        "released": released,
        "counts": summarize_ledger(ledger),
    }


def sorted_records(ledger: dict) -> list[dict]:
    return sorted(ledger.get("records", {}).values(), key=lambda item: item.get("row_number", 0))


def assess_final_images(ledger: dict) -> dict[str, list]:
    missing_generated = []
    image_validation = []
    for record in sorted_records(ledger):
        ok, error = validate_final_image(Path(record["generated_image_path"]))
        image_validation.append({"local_id": record["local_id"], "ok": ok, "error": error})
        if not ok:
            missing_generated.append(record["local_id"])
    return {"missing_generated": missing_generated, "image_validation": image_validation}


def write_output_workbook(records: list[dict], output_xlsx: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "main_images"
    headers = ["本地ID", "产品标题", "产品图片1", "主图"]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    link_font = Font(color="0563C1", underline="single")
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font

    for item_data in records:
        generated_path = Path(item_data["generated_image_path"])
        ws.append(
            [
                item_data["local_id"],
                item_data["title"],
                item_data["source_image_url"],
                str(generated_path.resolve()),
            ]
        )
        current_row = ws.max_row
        for col in (3, 4):
            cell = ws.cell(row=current_row, column=col)
            cell.hyperlink = cell.value
            cell.font = link_font

    set_widths(ws, [24, 80, 70, 90])
    ws.freeze_panes = "A2"
    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_xlsx)


def build(
    input_path: Path,
    limit: int | None,
    output_path: Path | None,
    output_root: Path | None,
    run_mode: str | None = None,
    allow_partial: bool = False,
    batch_id: str | None = None,
) -> dict:
    ledger, paths = load_or_prepare(input_path, limit, output_path, output_root, run_mode, batch_id)
    records = sorted_records(ledger)
    image_assessment = assess_final_images(ledger)
    counts = summarize_ledger(ledger)
    unfinished = unfinished_records(ledger.get("records", {}))
    if not allow_partial and (unfinished or image_assessment["missing_generated"]):
        raise SystemExit(
            "Strict build refused: unfinished rows or missing/invalid generated images remain. "
            "Use --allow-partial only for interim workbooks."
        )
    for record in records:
        if record.get("status") == "verified":
            normalize_generated_image(Path(record["generated_image_path"]))
    output_xlsx = paths["output_xlsx"]
    write_output_workbook(records, output_xlsx)
    payload = {
        "input_xlsx": str(input_path.resolve()),
        "batch_id": paths["batch_id"],
        "batch_dir": str(paths["base_dir"].resolve()),
        "run_mode": ledger.get("run_mode"),
        "requested_limit": ledger.get("requested_limit"),
        "valid_row_count": ledger.get("valid_row_count"),
        "items": records,
        "skips": ledger.get("skips", []),
        "originals_dir": str(paths["originals_dir"].resolve()),
        "generated_dir": str(paths["generated_dir"].resolve()),
        "generated_fast_dir": str(paths["generated_fast_dir"].resolve()),
        "staging_dir": str(paths["staging_dir"].resolve()),
        "review_queue_dir": str(paths["review_queue_dir"].resolve()),
        "output_xlsx": str(output_xlsx.resolve()),
        "status_ledger": str(paths["status_path"].resolve()),
        "status_counts": counts,
        "missing_generated": image_assessment["missing_generated"],
        "image_validation": image_assessment["image_validation"],
        "manifest_path": str(paths["manifest_path"].resolve()),
        "log_path": str(paths["log_path"].resolve()),
    }
    write_json(
        paths["log_path"],
        {
            "skips": payload["skips"],
            "missing_generated": payload["missing_generated"],
            "image_validation": payload["image_validation"],
        },
    )
    return payload


def verify(
    input_path: Path,
    limit: int | None,
    output_path: Path | None,
    output_root: Path | None,
    run_mode: str | None = None,
    batch_id: str | None = None,
) -> dict:
    ledger, paths = load_or_prepare(input_path, limit, output_path, output_root, run_mode, batch_id)
    records = sorted_records(ledger)
    counts = summarize_ledger(ledger)
    checks: dict[str, object] = {}
    unfinished = unfinished_records(ledger.get("records", {}))
    if unfinished:
        checks["unfinished_rows"] = [record["local_id"] for record in unfinished]

    image_assessment = assess_final_images(ledger)
    if image_assessment["missing_generated"]:
        checks["missing_or_invalid_images"] = image_assessment["missing_generated"]

    if ledger.get("valid_row_count") != len(records):
        checks["row_count_mismatch"] = {
            "valid_row_count": ledger.get("valid_row_count"),
            "ledger_records": len(records),
        }

    output_xlsx = paths["output_xlsx"]
    if not output_xlsx.exists():
        checks["missing_output_workbook"] = str(output_xlsx.resolve())
    else:
        try:
            wb = load_workbook(output_xlsx, read_only=True, data_only=True)
            ws = wb.active
            workbook_rows = max(ws.max_row - 1, 0)
            path_values = [clean_cell(ws.cell(row=row, column=4).value) for row in range(2, ws.max_row + 1)]
            wb.close()
            expected_paths = [str(Path(record["generated_image_path"]).resolve()) for record in records]
            if workbook_rows != len(records):
                checks["workbook_row_count_mismatch"] = {
                    "workbook_rows": workbook_rows,
                    "ledger_records": len(records),
                }
            if path_values != expected_paths:
                checks["workbook_main_image_paths_mismatch"] = True
        except Exception as exc:  # noqa: BLE001 - expose workbook verification issue.
            checks["workbook_error"] = str(exc)

    return {
        "ok": not checks,
        "status_ledger": str(paths["status_path"].resolve()),
        "manifest_path": str(paths["manifest_path"].resolve()),
        "output_xlsx": str(paths["output_xlsx"].resolve()),
        "batch_id": paths["batch_id"],
        "batch_dir": str(paths["base_dir"].resolve()),
        "run_mode": ledger.get("run_mode"),
        "requested_limit": ledger.get("requested_limit"),
        "valid_row_count": ledger.get("valid_row_count"),
        "counts": counts,
        "checks": checks,
        "image_validation": image_assessment["image_validation"],
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Prepare/build Mango Excel main-image batches.")
    subparsers = parser.add_subparsers(dest="command", required=True)
    for command in ("prepare", "build", "status", "next", "mark", "verify", "claim", "release"):
        sub = subparsers.add_parser(command)
        sub.add_argument("--input", required=True, type=Path, help="Input .xlsx file")
        sub.add_argument("--limit", type=int, default=3, help="Number of valid products to process; use 0 for all")
        sub.add_argument("--output", type=Path, default=None, help="Output .xlsx path")
        sub.add_argument("--output-root", type=Path, default=None, help="Root folder for batch outputs")
        sub.add_argument("--run-mode", choices=("sample", "full"), default=None, help="Queue mode; inferred from --limit when omitted")
        sub.add_argument("--batch-id", default=None, help="Specific batch folder id; defaults to latest batch when available")
        if command == "prepare":
            sub.add_argument("--new-task", action="store_true", help="Create a fresh timestamped batch folder and make it the latest batch")
        if command in ("next", "claim"):
            sub.add_argument("--count", type=int, default=20, help="Number of unfinished products to return")
        if command == "build":
            sub.add_argument("--allow-partial", action="store_true", help="Allow interim workbook with unfinished/missing images")
        if command == "mark":
            sub.add_argument("--local-id", required=True, help="Product local ID to update")
            sub.add_argument("--status", required=True, choices=ALL_STATUSES, help="New status")
            sub.add_argument("--error", default=None, help="Failure/error note")
            sub.add_argument("--owner", default=None, help="Worker owner that claimed the row")
        if command == "claim":
            sub.add_argument("--owner", required=True, help="Worker owner for claimed rows")
            sub.add_argument("--lease-minutes", type=int, default=30, help="Claim lease duration")
        if command == "release":
            sub.add_argument("--owner", required=True, help="Worker owner whose rows should be released")
            sub.add_argument("--local-id", action="append", default=None, help="Specific local ID to release; repeatable")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    input_path = args.input.resolve()
    if input_path.suffix.lower() != ".xlsx":
        raise SystemExit("Input must be a .xlsx file.")
    if not input_path.exists():
        raise SystemExit(f"Input not found: {input_path}")

    limit = None if args.limit == 0 else args.limit
    output_root = args.output_root.resolve() if args.output_root else None

    if args.command == "prepare":
        result = prepare(input_path, limit, args.output, output_root, args.run_mode, args.batch_id, args.new_task)
    elif args.command == "build":
        result = build(input_path, limit, args.output, output_root, args.run_mode, args.allow_partial, args.batch_id)
    elif args.command == "status":
        result = status(input_path, limit, args.output, output_root, args.run_mode, args.batch_id)
    elif args.command == "next":
        result = next_items(input_path, limit, args.output, output_root, args.count, args.run_mode, args.batch_id)
    elif args.command == "verify":
        result = verify(input_path, limit, args.output, output_root, args.run_mode, args.batch_id)
    elif args.command == "claim":
        result = claim_items(input_path, limit, args.output, output_root, args.count, args.owner, args.lease_minutes, args.run_mode, args.batch_id)
    elif args.command == "release":
        result = release_items(input_path, limit, args.output, output_root, args.owner, args.local_id, args.run_mode, args.batch_id)
    else:
        result = mark_item(input_path, limit, args.output, output_root, args.local_id, args.status, args.error, args.owner, args.run_mode, args.batch_id)
    print(json.dumps(result, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    sys.exit(main())
