from __future__ import annotations

import importlib.util
import shutil
import subprocess
import sys
import unittest
import uuid
from pathlib import Path

from openpyxl import load_workbook, Workbook
from PIL import Image


ROOT = Path(__file__).resolve().parents[1]
SCRIPT_PATH = ROOT / "scripts" / "process_mango_excel.py"
COMPOSE_SCRIPT_PATH = ROOT / "scripts" / "compose_batch_images.py"


def load_process_module():
    module_name = "process_mango_excel_under_test"
    spec = importlib.util.spec_from_file_location(module_name, SCRIPT_PATH)
    module = importlib.util.module_from_spec(spec)
    assert spec.loader is not None
    sys.modules[module_name] = module
    spec.loader.exec_module(module)
    return module


class MangoBatchTests(unittest.TestCase):
    def setUp(self) -> None:
        temp_parent = ROOT / "tmp" / "tests"
        temp_parent.mkdir(parents=True, exist_ok=True)
        self.root = temp_parent / f"case-{uuid.uuid4().hex}"
        self.root.mkdir()
        self.proc = load_process_module()
        self.input_xlsx = self.root / "sample_5_rows.xlsx"
        self.output_root = self.root / "out"
        self.source_image = self.root / "source.png"
        self._write_source_image()
        self._write_workbook()

    def tearDown(self) -> None:
        shutil.rmtree(self.root, ignore_errors=True)

    def _write_source_image(self) -> None:
        image = Image.new("RGB", (240, 180), "#f4f6f8")
        image.save(self.source_image)

    def _write_workbook(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws.append(["ignored row"])
        ws.append(["\u672c\u5730ID", "\u4ea7\u54c1\u6807\u9898", "\u4ea7\u54c1\u56fe\u72471"])
        for idx in range(1, 6):
            ws.append([f"ID{idx:03d}", f"Auto part {idx}", self.source_image.as_uri()])
        wb.save(self.input_xlsx)

    def _make_final_image(self, local_id: str, batch_id: str | None = None) -> None:
        paths = self.proc.workbook_paths(self.input_xlsx, output_root=self.output_root, batch_id=batch_id)
        final_path = paths["generated_dir"] / f"{local_id}.png"
        final_path.parent.mkdir(parents=True, exist_ok=True)
        Image.new("RGB", (800, 800), "#ffffff").save(final_path)

    def test_full_prepare_replaces_sample_queue(self) -> None:
        sample = self.proc.prepare(
            self.input_xlsx,
            limit=3,
            output_path=None,
            output_root=self.output_root,
            run_mode="sample",
        )
        self.assertEqual(sample["valid_row_count"], 5)
        self.assertEqual(sample["status_counts"]["total"], 3)

        full = self.proc.prepare(
            self.input_xlsx,
            limit=None,
            output_path=None,
            output_root=self.output_root,
            run_mode="full",
        )

        self.assertEqual(full["run_mode"], "full")
        self.assertEqual(full["requested_limit"], 0)
        self.assertEqual(full["valid_row_count"], 5)
        self.assertEqual(full["status_counts"]["total"], 5)

    def test_strict_build_fails_when_images_are_missing(self) -> None:
        self.proc.prepare(self.input_xlsx, None, None, self.output_root, run_mode="full")

        with self.assertRaises(SystemExit):
            self.proc.build(self.input_xlsx, None, None, self.output_root, run_mode="full", allow_partial=False)

        partial = self.proc.build(
            self.input_xlsx,
            None,
            None,
            self.output_root,
            run_mode="full",
            allow_partial=True,
        )
        self.assertEqual(len(partial["missing_generated"]), 5)
        self.assertTrue(Path(partial["output_xlsx"]).exists())

    def test_verify_requires_all_rows_verified_and_workbook_complete(self) -> None:
        self.proc.prepare(self.input_xlsx, None, None, self.output_root, run_mode="full")
        for idx in range(1, 6):
            self._make_final_image(f"ID{idx:03d}")

        failed = self.proc.verify(self.input_xlsx, None, None, self.output_root, run_mode="full")
        self.assertFalse(failed["ok"])
        self.assertIn("unfinished_rows", failed["checks"])

        for idx in range(1, 6):
            self.proc.mark_item(
                self.input_xlsx,
                None,
                None,
                self.output_root,
                f"ID{idx:03d}",
                "verified",
                None,
                owner=None,
                run_mode="full",
            )
        built = self.proc.build(self.input_xlsx, None, None, self.output_root, run_mode="full", allow_partial=False)
        self.assertTrue(Path(built["output_xlsx"]).exists())
        passed = self.proc.verify(self.input_xlsx, None, None, self.output_root, run_mode="full")
        self.assertTrue(passed["ok"], passed)

    def test_claims_are_disjoint_and_expired_claims_are_reclaimed(self) -> None:
        self.proc.prepare(self.input_xlsx, None, None, self.output_root, run_mode="full")

        first = self.proc.claim_items(
            self.input_xlsx,
            None,
            None,
            self.output_root,
            count=2,
            owner="worker-a",
            lease_minutes=30,
            run_mode="full",
        )
        second = self.proc.claim_items(
            self.input_xlsx,
            None,
            None,
            self.output_root,
            count=2,
            owner="worker-b",
            lease_minutes=30,
            run_mode="full",
        )
        first_ids = {item["local_id"] for item in first["items"]}
        second_ids = {item["local_id"] for item in second["items"]}
        self.assertEqual(len(first_ids), 2)
        self.assertEqual(len(second_ids), 2)
        self.assertFalse(first_ids & second_ids)

        ledger_path = Path(first["status_ledger"])
        ledger = self.proc.read_json(ledger_path, {})
        for local_id in first_ids:
            ledger["records"][local_id]["lease_expires_at"] = "2000-01-01T00:00:00+00:00"
        self.proc.write_json(ledger_path, ledger)

        reclaimed = self.proc.claim_items(
            self.input_xlsx,
            None,
            None,
            self.output_root,
            count=2,
            owner="worker-c",
            lease_minutes=30,
            run_mode="full",
        )
        reclaimed_ids = {item["local_id"] for item in reclaimed["items"]}
        self.assertTrue(reclaimed_ids & first_ids)

    def test_release_returns_owned_claims_to_pending(self) -> None:
        self.proc.prepare(self.input_xlsx, None, None, self.output_root, run_mode="full")
        claimed = self.proc.claim_items(
            self.input_xlsx,
            None,
            None,
            self.output_root,
            count=2,
            owner="worker-a",
            lease_minutes=30,
            run_mode="full",
        )
        released = self.proc.release_items(
            self.input_xlsx,
            None,
            None,
            self.output_root,
            owner="worker-a",
            run_mode="full",
        )
        self.assertEqual(released["released"], 2)
        self.assertEqual(released["counts"].get("pending"), 5)

    def test_new_task_creates_fresh_batch_folder_and_updates_latest_pointer(self) -> None:
        first = self.proc.prepare(
            self.input_xlsx,
            None,
            None,
            self.output_root,
            run_mode="full",
            new_task=True,
        )
        self._make_final_image("ID001", batch_id=first["batch_id"])
        self.proc.mark_item(
            self.input_xlsx,
            None,
            None,
            self.output_root,
            "ID001",
            "verified",
            None,
            owner=None,
            run_mode="full",
            batch_id=first["batch_id"],
        )

        second = self.proc.prepare(
            self.input_xlsx,
            None,
            None,
            self.output_root,
            run_mode="full",
            new_task=True,
        )

        self.assertNotEqual(first["batch_id"], second["batch_id"])
        self.assertNotEqual(first["batch_dir"], second["batch_dir"])
        self.assertEqual(second["status_counts"].get("pending"), 5)
        self.assertFalse(Path(second["items"][0]["generated_image_path"]).exists())

        latest = self.proc.status(self.input_xlsx, None, None, self.output_root, run_mode="full")
        self.assertEqual(latest["batch_id"], second["batch_id"])

    def test_template_fallback_uses_generated_fast_and_template_status(self) -> None:
        prepared = self.proc.prepare(self.input_xlsx, None, None, self.output_root, run_mode="full")
        subprocess.run(
            [
                sys.executable,
                str(COMPOSE_SCRIPT_PATH),
                "--manifest",
                prepared["manifest_path"],
                "--status-ledger",
                prepared["status_ledger"],
                "--limit",
                "1",
            ],
            check=True,
            capture_output=True,
            text=True,
        )

        ledger = self.proc.read_json(Path(prepared["status_ledger"]), {})
        record = ledger["records"]["ID001"]
        self.assertEqual(record["status"], "template_generated")
        self.assertFalse(Path(record["generated_image_path"]).exists())
        self.assertTrue(Path(record["generated_fast_image_path"]).exists())

    def test_final_workbook_row_count_matches_verified_records(self) -> None:
        self.proc.prepare(self.input_xlsx, None, None, self.output_root, run_mode="full")
        for idx in range(1, 6):
            local_id = f"ID{idx:03d}"
            self._make_final_image(local_id)
            self.proc.mark_item(
                self.input_xlsx,
                None,
                None,
                self.output_root,
                local_id,
                "verified",
                None,
                owner=None,
                run_mode="full",
            )
        built = self.proc.build(self.input_xlsx, None, None, self.output_root, run_mode="full", allow_partial=False)
        wb = load_workbook(built["output_xlsx"], read_only=True)
        try:
            ws = wb.active
            self.assertEqual(ws.max_row - 1, 5)
        finally:
            wb.close()


if __name__ == "__main__":
    unittest.main()
